import csv
import os
import time
import json
import base64
import urllib.request
import urllib.parse
import sys
import tkinter as tk
from tkinter import filedialog

# Tkinter Setup (for file dialogs only)
tk_root = tk.Tk()
tk_root.withdraw()

# Robust script directory detection for Resolve Environment
if "__file__" in globals() and __file__:
    script_dir = os.path.dirname(os.path.abspath(__file__))
else:
    # Fallback paths for different platforms when __file__ is not defined
    if sys.platform == "win32":
        script_dir = r"c:\ProgramData\Blackmagic Design\DaVinci Resolve\Fusion\Scripts\Utility"
    elif sys.platform == "darwin":
        script_dir = "/Library/Application Support/Blackmagic Design/DaVinci Resolve/Fusion/Scripts/Utility"
    else:
        # Linux
        script_dir = "/opt/resolve/Fusion/Scripts/Utility"

# Cross-platform library path detection
# We check both the script directory and its parent directory (for when it's in a subdirectory)
parent_dir = os.path.dirname(script_dir)
extra_paths = [
    os.path.join(script_dir, "lib"),
    os.path.join(parent_dir, "lib")
]
if sys.platform == "win32":
    extra_paths.append(os.path.join(script_dir, ".venv", "Lib", "site-packages"))
    extra_paths.append(os.path.join(parent_dir, ".venv", "Lib", "site-packages"))
else:
    # macOS/Linux .venv structure: lib/pythonX.X/site-packages
    for base in [script_dir, parent_dir]:
        venv_lib = os.path.join(base, ".venv", "lib")
        if os.path.exists(venv_lib):
            for d in os.listdir(venv_lib):
                if d.startswith("python"):
                    extra_paths.append(os.path.join(venv_lib, d, "site-packages"))

for p in extra_paths:
    if os.path.exists(p) and p not in sys.path:
        sys.path.insert(0, p)

try:
    import DaVinciResolveScript as bmd
    fu = bmd.scriptapp("Fusion")
    resolve = bmd.scriptapp("Resolve")
    ui = fu.UIManager
    disp = bmd.UIDispatcher(ui)
except ImportError:
    # use davinci resolve fusion app directly
    fu = resolve.Fusion()
    ui = fu.UIManager
    disp = bmd.UIDispatcher(ui)

# --- Platform-Safe File Dialogs ---
def request_file(title="Select File", mode="open", ext_filter="CSV Files (*.csv)", default_ext=".csv"):
    """
    Cross-platform file dialog using tkinter
    mode: "open", "save", or "directory"
    """
    # Create extension filter
    if "CSV" in ext_filter:
        filetypes = [("CSV Files", "*.csv"), ("All Files", "*.*")]
    else:
        filetypes = [("All Files", "*.*")]
        
    tk_root.update()
    tk_root.attributes("-topmost", True)
    
    try:
        if mode == "directory":
            path = filedialog.askdirectory(parent=tk_root, title=title)
        elif mode == "save":
            path = filedialog.asksaveasfilename(parent=tk_root, title=title, defaultextension=default_ext, filetypes=filetypes)
        else:
            path = filedialog.askopenfilename(parent=tk_root, title=title, defaultextension=default_ext, filetypes=filetypes)
    finally:
        tk_root.attributes("-topmost", False)
        
    return path if path else None

def request_yes_no(title, message):
    """Simple yes/no dialog using Fusion UI."""
    win_id = "ConfirmDialog"
    dialog_layout = ui.VGroup({"Spacing": 10}, [
        ui.Label({"Text": message, "WordWrap": True, "Alignment": {"AlignHCenter": True}}),
        ui.HGroup({"Weight": 0}, [
            ui.Button({"ID": "y_btn", "Text": "Yes"}),
            ui.Button({"ID": "n_btn", "Text": "No"}),
        ])
    ])
    
    dlg = disp.AddWindow({"WindowTitle": title, "ID": win_id, "Geometry": [400, 400, 300, 150]}, dialog_layout)
    result = {"value": False}
    
    def on_yes(ev):
        result["value"] = True
        disp.ExitLoop()
    def on_no(ev):
        result["value"] = False
        disp.ExitLoop()
        
    dlg.On["y_btn"].Clicked = on_yes
    dlg.On["n_btn"].Clicked = on_no
    dlg.On[win_id].Close = on_no
    
    dlg.Show()
    disp.RunLoop()
    dlg.Hide()
    return result["value"]

try:
    import cv2
    import numpy as np
    HAS_OPENCV = True
except ImportError as e:
    print("OpenCV or NumPy not found. Attempting automatic installation...")
    try:
        import subprocess
        lib_dir = os.path.join(script_dir, "lib")
        os.makedirs(lib_dir, exist_ok=True)
        
        # Run pip install targeting the lib_dir
        subprocess.check_call([
            sys.executable, "-m", "pip", "install",
            "--target", lib_dir,
            "numpy>=1.20.0",
            "opencv-python-headless>=4.5.0"
        ])
        
        # Add to sys.path if not already present
        if lib_dir not in sys.path:
            sys.path.insert(0, lib_dir)
            
        import cv2
        import numpy as np
        HAS_OPENCV = True
        print("Successfully installed and loaded OpenCV and NumPy!")
    except Exception as install_err:
        HAS_OPENCV = False
        print(f"Warning: OpenCV (cv2) or NumPy could not be loaded/installed: {install_err}. Thumbnail resizing will be skipped.")

def resize_image_opencv(image_path, target_width=720):
    """Resizes an image to target_width if it exceeds it, maintaining aspect ratio."""
    if not HAS_OPENCV:
        return False
        
    try:
        # OpenCV doesn't handle unicode paths well on Windows sometimes, 
        # but for standard paths it should be fine.
        img = cv2.imread(image_path)
        if img is None:
            return False
            
        h, w = img.shape[:2]
        if w <= target_width:
            return True
            
        scale = target_width / w
        new_h = int(h * scale)
        
        # Use INTER_AREA for downsampling (best quality)
        resized = cv2.resize(img, (target_width, new_h), interpolation=cv2.INTER_AREA)
        cv2.imwrite(image_path, resized)
        return True
    except Exception as e:
        print(f"Error resizing {image_path}: {e}")
        return False

def save_clip_thumbnail_to_file(timeline, start_tc, target_path):
    """Switches to color page, moves playhead, gets thumbnail, and writes to target_path using OpenCV."""
    resolve.OpenPage("color")
    timeline.SetCurrentTimecode(start_tc)
    time.sleep(1)
    
    thumb_data = timeline.GetCurrentClipThumbnailImage()
    if not thumb_data or not thumb_data.get("data"):
        print(f"Failed to retrieve thumbnail at {start_tc}")
        return False
        
    if not HAS_OPENCV:
        print("OpenCV is not loaded; cannot process thumbnail.")
        return False
        
    try:
        width = thumb_data["width"]
        height = thumb_data["height"]
        img_bytes = base64.b64decode(thumb_data["data"])
        nparr = np.frombuffer(img_bytes, np.uint8)
        cvimg = nparr.reshape((height, width, 3))
        cvimg = cv2.cvtColor(cvimg, cv2.COLOR_RGB2BGR)
        cv2.imwrite(target_path, cvimg)
        return True
    except Exception as e:
        print(f"Error generating thumbnail at {start_tc}: {e}")
        return False

def frame_to_timecode(frame_number: int, fps=24.0):
    """
    Convert a frame number to timecode (HH:MM:SS:FF)

    Parameters:
        frame_number (int): The frame number to convert.
        fps (float): Frames per second.

    Returns:
        str: Timecode in the format HH:MM:SS:FF
    """
    # Use rounded integer FPS for standard non-drop frame timecode counting
    fps_int = int(round(float(fps)))
    frame_number = int(frame_number)
    
    total_seconds = frame_number // fps_int
    frames = frame_number % fps_int

    hours = total_seconds // 3600
    minutes = (total_seconds % 3600) // 60
    seconds = total_seconds % 60

    return f"{hours:02d}:{minutes:02d}:{seconds:02d}:{frames:02d}"

def timecode_to_frame(tc: str, fps=24.0):
    fps_int = int(round(float(fps)))
    parts = tc.replace(';', ':').split(':')
    if len(parts) != 4: return 0
    h, m, s, f = [int(p) for p in parts]
    return (h * 3600 + m * 60 + s) * fps_int + f



class ShotRecord:
    def __init__(self, shot_id="", name="", thumb="", dur="", start="", end="", status="", assign="", diff="", prior="", notes="", timeline=""):
        self.shot_id = shot_id
        self.name = name
        self.thumb = thumb
        self.duration = dur
        self.start_tc = start
        self.end_tc = end
        self.status = status
        self.assign = assign
        self.difficulty = diff
        self.prior = prior
        self.notes = notes
        self.timeline = timeline

    def to_list(self):
        return [self.shot_id, self.name, self.thumb, self.duration, self.start_tc, self.end_tc, self.status, self.assign, self.difficulty, self.prior, self.notes, self.timeline]

    @classmethod
    def from_list(cls, row):
        while len(row) < 12: row.append("")
        return cls(*row[:12])


def update_shot_vfx_marker(shot):
    """Creates or updates a VFX marker on the appropriate Resolve timeline for the given shot."""
    try:
        projectManager = resolve.GetProjectManager()
        project = projectManager.GetCurrentProject()
        if not project:
            return False
            
        final_timeline = shot.timeline
        tl = get_timeline_by_name(project, final_timeline)
        if not tl:
            print(f"Warning: Timeline '{final_timeline}' not found. Cannot create/update VFX marker.")
            return False
            
        try:
            framerate_str = tl.GetSetting("timelineFrameRate")
            fps = float(framerate_str) if framerate_str else 24.0
        except:
            fps = 24.0
        
        absolute_frame = timecode_to_frame(shot.start_tc, fps)
        timeline_start_frame = tl.GetStartFrame()
        frame_id = absolute_frame - timeline_start_frame
        
        final_id = shot.shot_id
        if not final_id:
            return False
            
        marker_name = f"XS_{final_id}"
        marker_color = "Green" if shot.status and shot.status.lower() == "approved" else "Lemon"
        
        keywords = []
        if shot.prior:
            keywords.append(f"P{shot.prior}")
        if shot.assign:
            keywords.append(f"Assign: {shot.assign}")
        if shot.difficulty:
            keywords.append(f"Difficulty: {shot.difficulty}")
        
        keyword_line = f"{'\n'.join(keywords)}" if keywords else ""
        
        if shot.notes and keyword_line:
            marker_note = f"{keyword_line}\n{shot.notes}"
        elif shot.notes:
            marker_note = shot.notes
        else:
            marker_note = keyword_line
        
        # Delete existing marker at this frame or with the same name
        markers = tl.GetMarkers()
        if markers:
            for f_id, m_info in list(markers.items()):
                if m_info['name'] == marker_name or int(f_id) == frame_id:
                    tl.DeleteMarkerAtFrame(int(f_id))
                    print(f"Deleted existing marker '{m_info['name']}' at frame {f_id}.")
        
        # Create Marker
        success = tl.AddMarker(frame_id, marker_color, marker_name, marker_note, 1)
        if success:
            print(f"Added/Updated marker '{marker_name}' at {shot.start_tc} on timeline '{final_timeline}'.")
            return True
        else:
            print(f"Failed to add marker '{marker_name}' at {shot.start_tc} on timeline '{final_timeline}'.")
            return False
    except Exception as ex:
        print(f"Failed to create VFX marker: {ex}")
        return False


# Global state for popup dialog
global_seq_name = "SEQ"
global_shot_num = 10
global_proj_name = "PROJ"
global_version = "v0001"
master_data = [] # List of ShotRecord objects
current_session_file = None
DEFAULT_STATUSES = ["Pending", "WIP", "Approved", "Final", "Omit"]

# Define the UI layout
layout = ui.VGroup({"Spacing": 10}, [
    # Session Bar
    ui.HGroup({"Spacing": 10, "Weight": 0}, [
        ui.Label({"Text": "Session File:", "Weight": 0}),
        ui.LineEdit({"ID": "session_path", "ReadOnly": True, "Text": "None (Work will not be saved)", "Weight": 1}),
        ui.Button({"ID": "link_session_btn", "Text": "Link...", "Weight": 0}),
    ]),

    ui.HGroup({"Spacing": 10, "Weight": 0}, [
        ui.Button({"ID": "refresh_btn", "Text": "Add Current Timeline Clip", "Weight": 2}),
        ui.Button({"ID": "edit_btn", "Text": "Edit Selected", "Weight": 1}),
        ui.Button({"ID": "batch_edit_btn", "Text": "Batch Edit...", "Weight": 1}),
        ui.Button({"ID": "delete_btn", "Text": "Delete Selected", "Weight": 1}),
    ]),
    
    # Filter Bar
    ui.HGroup({"Spacing": 10, "Weight": 0}, [
        ui.Label({"Text": "Filter:", "Weight": 0}),
        ui.ComboBox({"ID": "filter_col", "Weight": 1}),
        ui.LineEdit({"ID": "filter_text", "PlaceholderText": "Search...", "Weight": 2}),
        ui.Button({"ID": "clear_filter_btn", "Text": "Clear", "Weight": 0}),
    ]),

    ui.Tree({
        "ID": "clips_tree",
        "ColumnCount": 12,
        "AlternatingRowColors": True,
        "SortingEnabled": True,
        "Weight": 1,
        "HeaderHidden": False,
        "SelectionMode": "ExtendedSelection",
        'Events' : { 'ItemDoubleClicked' : True }
    }),
    
    ui.HGroup({"Spacing": 10, "Weight": 0}, [
        ui.Button({"ID": "import_csv_btn", "Text": "Import CSV", "Weight": 1}),
        ui.Button({"ID": "export_csv_btn", "Text": "Export CSV", "Weight": 1}),
        ui.Button({"ID": "export_excel_btn", "Text": "Export Excel", "Weight": 1}),
        ui.Button({"ID": "export_thumb_btn", "Text": "Export Thumbnails", "Weight": 1}),
        ui.Button({"ID": "export_clips_btn", "Text": "Export Clips", "Weight": 1}),
    ]),

    # ui.HGroup({"Spacing": 10, "Weight": 0}, [
    #     ui.Label({"Text": "GSheet URL:", "Weight": 0}),
    #     ui.LineEdit({"ID": "gsheet_url", "PlaceholderText": "Paste Web App URL here...", "Weight": 1}),
    #     ui.Button({"ID": "sync_gsheet_btn", "Text": "Sync to Google Sheet", "Weight": 0, "StyleSheet": "background-color: #2e7d32; color: white; font-weight: bold;"}),
    # ]),
    ui.HGroup({"Weight": 0}, [
        ui.Label({"Text": "Export Scope (CSV/Excel/Thumb/Clips):", "Weight": 0}),
        ui.ComboBox({"ID": "export_scope_combo", "Weight": 1}),
    ])
])

win = disp.AddWindow({"WindowTitle": "Timeline Shot Manager", "ID": "MainWindow", "Geometry": [100, 150, 900, 600]}, layout)

scope_combo = win.Find("export_scope_combo")
scope_combo.AddItem("All Items")
scope_combo.AddItem("Filtered Items")
scope_combo.AddItem("Selected Items")
scope_combo.CurrentIndex = 0

tree = win.Find("clips_tree")
tree.SetHeaderLabels(["Shot ID", "Clip Name", "", "Duration", "Start TC", "End TC", "Status", "Assign", "Difficulty", "Prior", "Notes", "Timeline"])

tree.ColumnWidth[0] = 120 # Shot ID
tree.ColumnWidth[1] = 200 # Clip Name
tree.ColumnWidth[2] = 0 # Thumbnails
tree.ColumnWidth[3] = 80  # Duration
tree.ColumnWidth[4] = 100 # Start TC
tree.ColumnWidth[5] = 100 # End TC
tree.ColumnWidth[6] = 100 # Status
tree.ColumnWidth[7] = 100 # Assign
tree.ColumnWidth[8] = 100 # Difficulty
tree.ColumnWidth[9] = 100 # Prior
tree.ColumnWidth[10] = 200 # Notes
tree.ColumnWidth[11] = 0  # Timeline (Hidden)

# Setup filter columns
filter_combo = win.Find("filter_col")
header_labels = ["Shot ID", "Clip Name", "", "Duration", "Start TC", "End TC", "Status", "Assign", "Difficulty", "Prior", "Notes", "Timeline"]
filter_indices = []

# Add "All Columns" option as the default
filter_combo.AddItem("All Columns")
filter_indices.append(-1) # -1 representing "All Columns"

for i, label in enumerate(header_labels):
    if label: # Don't add empty labels (hidden columns) to the filter
        filter_combo.AddItem(label)
        filter_indices.append(i)
filter_combo.CurrentIndex = 0 # Default to All Columns


# Cache clips directly to map tree items to clip objects if needed
clip_cache = []
fps = 24

def get_project_name():
    """Returns the current Resolve project name."""
    projectManager = resolve.GetProjectManager()
    project = projectManager.GetCurrentProject()
    return project.GetName() if project else None

def get_timeline_by_name(project, name):
    """Returns the timeline object matching the given name, or None."""
    if not project or not name:
        return None
    count = project.GetTimelineCount()
    for i in range(1, int(count) + 1):
        t = project.GetTimelineByIndex(i)
        if t.GetName() == name:
            return t
    return None

def auto_save_session():
    """Saves the master_data and global UI settings to the currently linked session file."""
    global current_session_file, global_seq_name, global_shot_num
    if not current_session_file:
        return
        
    try:
        # Save CSV Data
        dir_path = os.path.dirname(current_session_file)
        if dir_path and not os.path.exists(dir_path):
            os.makedirs(dir_path)
            
        with open(current_session_file, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(["Shot ID", "Clip Name", "Thumbnails", "Duration", "Start TC", "End TC", "Status", "Assign", "Difficulty", "Prior", "Notes", "Timeline"])
            for shot in master_data:
                writer.writerow(shot.to_list())
        
        # Save Global Settings to Preferences
        proj_name = get_project_name()
        if proj_name:
            gsheet_field = win.Find("gsheet_url")
            gsheet_url = gsheet_field.Text.strip() if gsheet_field else ""
            sessions = fu.GetData("FxShotManager_Sessions") or {}
            sessions[proj_name] = {
                "path": current_session_file,
                "seq": global_seq_name,
                "shot": global_shot_num,
                "gProj": global_proj_name,
                "gVers": global_version,
                "gsheet": gsheet_url
            }
            fu.SetData("FxShotManager_Sessions", sessions)
            
        print(f"Auto-saved to {current_session_file} (Seq: {global_seq_name}, Next: {global_shot_num})")
    except Exception as e:
        print(f"Auto-save failed: {e}")

def refresh_tree():
    """Clears and re-populates the tree based on the current filter."""
    tree.Clear()
    
    filter_idx = win.Find("filter_col").CurrentIndex
    filter_text = win.Find("filter_text").Text.strip().lower()
    
    # Map combo index to actual ShotRecord column index
    real_idx = filter_indices[filter_idx] if filter_idx < len(filter_indices) else -1
    
    item_map = {}
    for shot in master_data:
        # Check filter
        if filter_text:
            if real_idx == -1:
                # Search all columns except the empty/hidden ones
                found = False
                shot_list = shot.to_list()
                for idx in filter_indices:
                    if idx != -1 and filter_text in str(shot_list[idx]).lower():
                        found = True
                        break
                if not found:
                    continue
            else:
                # Search specifically selected column
                val = str(shot.to_list()[real_idx]).lower()
                if filter_text not in val:
                    continue
        
        item = tree.NewItem()
        for c, val in enumerate(shot.to_list()):
            item.SetText(c, str(val))
        tree.AddTopLevelItem(item)
        item_map[shot] = item
        
    return item_map


def add_current_clip(ev):
    global fps, master_data
    projectManager = resolve.GetProjectManager()
    currentProject = projectManager.GetCurrentProject()
    if not currentProject:
        return
    currentTimeline = currentProject.GetCurrentTimeline()
    if not currentTimeline:
        return
        
    clip = currentTimeline.GetCurrentVideoItem()
    if not clip:
        print("No clip currently selected or under playhead.")
        return
        
    try:
        fps = float(currentTimeline.GetSetting('timelineFrameRate'))
    except:
        fps = 24.0
        
    shot = ShotRecord(
        shot_id="", 
        name=clip.GetName(),
        thumb="", 
        dur=str(clip.GetDuration()),
        start=frame_to_timecode(clip.GetStart(), fps),
        end=frame_to_timecode(clip.GetEnd(), fps),
        status="", 
        assign="", 
        diff="", 
        prior="", 
        notes="", 
        timeline=currentTimeline.GetName()
    )
    
    master_data.append(shot)
    
    # Check if we need to clear filter and refresh, or just append
    current_filter = win.Find("filter_text").Text.strip()
    if current_filter != "":
        win.On["filter_text"].TextChanged = None
        win.Find("filter_text").Text = ""
        win.On["filter_text"].TextChanged = on_filter_changed
        item_map = refresh_tree()
        auto_save_session()
        # Need to find the newly created TreeItem to open the dialog
        new_item = item_map.get(shot)
        if new_item:
            open_edit_dialog(new_item, is_new=True)
    else:
        # Just append to tree
        item = tree.NewItem()
        for c, val in enumerate(shot.to_list()):
            item.SetText(c, str(val))
        tree.AddTopLevelItem(item)
        auto_save_session()
        open_edit_dialog(item, is_new=True)


def open_edit_dialog(item, is_new=False):
    assign_history = set()
    for shot in master_data:
        if shot.assign and shot.assign.strip():
            assign_history.add(shot.assign.strip())
    assign_history = sorted(list(assign_history))

    # Safely capture the original signature of the item before showing the dialog
    item_text = item.Text if item.Text else {}
    orig_shot_id = item_text[0] if item_text else ""
    orig_name = item_text[1] if item_text else ""
    orig_start = item_text[4] if item_text else ""

    # Tracks if the user has fetched new clip data during this dialog session
    fetched_data = {
        "name": orig_name,
        "duration": item_text[3] if item_text else "",
        "start": orig_start,
        "end": item_text[5] if item_text else "",
        "status": item_text[6] if item_text else "",
        "assign": item_text[7] if item_text else "",
        "diff": item_text[8] if item_text else "",
        "prior": item_text[9] if item_text else "",
        "notes": item_text[10] if item_text else "",
        "timeline": item_text[11] if item_text else ""
    }

    dialog_layout = ui.VGroup({"Spacing": 10}, [
        ui.Button({"ID": "d_fetch_btn", "Text": "Update from Current Playhead", "Weight": 0}),
        ui.Label({"ID": "d_status_linked", "Text": f"Linked to: {fetched_data['name']} ({fetched_data['timeline']})", "Weight": 0, "Alignment": {"AlignHCenter": True}}),

        ui.Label({"Text": "Timeline", "Weight": 0}),
        ui.ComboBox({"ID": "d_timeline_combo", "Weight": 0}),
        
        ui.Label({"Text": "Final Shot ID", "Weight": 0}),
        ui.LineEdit({"ID": "d_shot_id", "Text": orig_shot_id if not is_new else f"{global_seq_name}_{global_shot_num:04d}", "Weight": 0}),

        ui.HGroup({"Spacing": 10, "Weight": 0}, [
            ui.VGroup({"Weight": 1}, [
                ui.Label({"Text": "Sequence Prefix", "Weight": 0}),
                ui.LineEdit({"ID": "d_seq_name", "Text": global_seq_name, "Weight": 0}),
            ]),
            ui.VGroup({"Weight": 1}, [
                ui.Label({"Text": "Shot Number", "Weight": 0}),
                ui.LineEdit({"ID": "d_shot_num", "Text": f"{global_shot_num:04d}" if is_new else (orig_shot_id.split('_')[-1] if '_' in orig_shot_id else f"{global_shot_num:04d}"), "Weight": 0}),
            ]),
        ]),

        ui.Label({"Text": "Status", "Weight": 0}),
        ui.HGroup({"Spacing": 5, "Weight": 0}, [
            ui.LineEdit({"ID": "d_status_line", "Text": fetched_data["status"], "Weight": 2}),
            ui.ComboBox({"ID": "d_status_combo", "Weight": 1}),
        ]),
        
        ui.Label({"Text": "Assign", "Weight": 0}),
        ui.HGroup({"Spacing": 5, "Weight": 0}, [
            ui.LineEdit({"ID": "d_assign_line", "Text": fetched_data["assign"], "Weight": 2}),
            ui.ComboBox({"ID": "d_assign_combo", "Weight": 1}),
        ]),

        ui.Label({"Text": "Difficulty", "Weight": 0}),
        ui.LineEdit({"ID": "d_difficulty", "Text": fetched_data["diff"], "Weight": 0}),
        
        ui.Label({"Text": "Prior", "Weight": 0}),
        ui.LineEdit({"ID": "d_prior", "Text": fetched_data["prior"], "Weight": 0}),

        ui.Label({"Text": "Notes", "Weight": 0}),
        ui.LineEdit({"ID": "d_notes", "Text": fetched_data["notes"],  'Alignment': { 'AlignTop' : True },'WordWrap': True, "Weight": 1}),
        
        ui.CheckBox({"ID": "d_create_marker", "Text": "Create VFX Marker", "Checked": True, "Weight": 0}),

        ui.HGroup({"Weight": 0}, [
            ui.Button({"ID": "d_ok", "Text": "OK"}),
            ui.Button({"ID": "d_cancel", "Text": "Cancel"}),
        ]),
        ui.VGap(1)
    ])
    
    dialog = disp.AddWindow({"WindowTitle": "Shot Details", "ID": "EditDialog", "Geometry": [300, 300, 500, 750]}, dialog_layout)
    
    combo = dialog.Find("d_assign_combo")
    combo.AddItem("Select Assign...")
    for p in assign_history:
        combo.AddItem(p)
        
    s_combo = dialog.Find("d_status_combo")
    s_combo.AddItem("Select Status...")
    # Gather unique statuses for history
    status_history = set(DEFAULT_STATUSES)
    for shot in master_data:
        if shot.status: status_history.add(shot.status)
    for s in sorted(list(status_history)):
        s_combo.AddItem(s)

    t_combo = dialog.Find("d_timeline_combo")
    t_names = []
    projectManager = resolve.GetProjectManager()
    project = projectManager.GetCurrentProject()
    if project:
        count = project.GetTimelineCount()
        for i in range(1, int(count) + 1):
            t_names.append(project.GetTimelineByIndex(i).GetName())
    t_names.sort()
    
    current_idx = 0
    for i, name in enumerate(t_names):
        t_combo.AddItem(name)
        if name == fetched_data["timeline"]:
            current_idx = i
    t_combo.CurrentIndex = current_idx

    def on_assign_select(ev):
        idx = combo.CurrentIndex
        if idx > 0:
            dialog.Find("d_assign_line").Text = combo.GetItemText(idx)
            
    dialog.On["d_assign_combo"].CurrentIndexChanged = on_assign_select

    def on_status_select(ev):
        idx = s_combo.CurrentIndex
        if idx > 0:
            dialog.Find("d_status_line").Text = s_combo.GetItemText(idx)
            
    dialog.On["d_status_combo"].CurrentIndexChanged = on_status_select
    
    def on_id_helper_changed(ev):
        seq = dialog.Find("d_seq_name").Text.strip()
        num_str = dialog.Find("d_shot_num").Text.strip()
        try:
            num = int(num_str)
            dialog.Find("d_shot_id").Text = f"{seq}_{num:04d}"
        except:
            dialog.Find("d_shot_id").Text = f"{seq}_{num_str}"
            
    dialog.On["d_seq_name"].TextChanged = on_id_helper_changed
    dialog.On["d_shot_num"].TextChanged = on_id_helper_changed

    def on_fetch_clicked(ev):
        projectManager = resolve.GetProjectManager()
        project = projectManager.GetCurrentProject()
        if not project: return
        timeline = project.GetCurrentTimeline()
        if not timeline: return
        clip = timeline.GetCurrentVideoItem()
        if not clip:
            print("No clip under playhead.")
            return

        try:
            f_fps = float(timeline.GetSetting('timelineFrameRate'))
        except:
            f_fps = 24.0

        fetched_data["name"] = clip.GetName()
        fetched_data["duration"] = str(clip.GetDuration())
        fetched_data["start"] = frame_to_timecode(clip.GetStart(), f_fps)
        fetched_data["end"] = frame_to_timecode(clip.GetEnd(), f_fps)
        fetched_data["timeline"] = timeline.GetName()
        
        dialog.Find("d_status_linked").Text = f"Updated to: {fetched_data['name']} ({fetched_data['timeline']})"
        print(f"Fetched new clip data: {fetched_data['name']} at {fetched_data['start']}")

    dialog.On["d_fetch_btn"].Clicked = on_fetch_clicked

    def on_ok(ev):
        global global_seq_name, global_shot_num
        final_id = dialog.Find("d_shot_id").Text.strip()
        seq_name = dialog.Find("d_seq_name").Text.strip()
        shot_num_str = dialog.Find("d_shot_num").Text.strip()
        assign_val = dialog.Find("d_assign_line").Text.strip()
        status_val = dialog.Find("d_status_line").Text.strip()
        notes = dialog.Find("d_notes").Text.strip()
        difficulty = dialog.Find("d_difficulty").Text.strip()
        prior = dialog.Find("d_prior").Text.strip()
        
        global_seq_name = seq_name
        
        # We need the original signature to find the record in master_data
        sig_name = orig_name
        sig_tc = orig_start

        # We manually apply the final_id to the item.
        # But we still try to update global_shot_num if it's a numeric ID for the NEXT add.
        try:
            num = int(shot_num_str)
            if is_new:
                global_shot_num = num + 10
        except ValueError:
            pass

        # Update Tree Item completely to avoid refresh_tree()
        item.SetText(0, final_id)
        item.SetText(1, fetched_data["name"])
        item.SetText(3, fetched_data["duration"])
        item.SetText(4, fetched_data["start"])
        item.SetText(5, fetched_data["end"])
        item.SetText(6, status_val)
        item.SetText(7, assign_val)
        item.SetText(8, difficulty)
        item.SetText(9, prior)
        item.SetText(10, notes)
        
        final_timeline = dialog.Find("d_timeline_combo").GetItemText(dialog.Find("d_timeline_combo").CurrentIndex)
        item.SetText(11, final_timeline)
        
        # Update Master Data
        for shot in master_data:
            if shot.name == sig_name and shot.start_tc == sig_tc:
                shot.shot_id = final_id
                shot.name = fetched_data["name"]
                shot.duration = fetched_data["duration"]
                shot.start_tc = fetched_data["start"]
                shot.end_tc = fetched_data["end"]
                shot.status = status_val
                shot.assign = assign_val
                shot.difficulty = difficulty
                shot.prior = prior
                shot.notes = notes
                shot.timeline = final_timeline
                
                # Create VFX Marker on timeline if checkbox is checked
                if dialog.Find("d_create_marker").Checked:
                    update_shot_vfx_marker(shot)
                break

        dialog.Hide()
        # refresh_tree() is no longer needed since we updated the item directly
        auto_save_session()

    def on_cancel(ev):
        dialog.Hide()

    dialog.On["d_ok"].Clicked = on_ok
    dialog.On["d_cancel"].Clicked = on_cancel
    dialog.On.EditDialog.Close = on_cancel
    dialog.Show()

def open_batch_edit_dialog(selected_items):
    # Field mapping: Name -> Index
    fields = {
        "status": 6,
        "assign": 7,
        "diff": 8,
        "prior": 9,
        "notes": 10,
        "timeline": 11
    }

    # Gather history for dropdowns
    assign_history = set()
    status_history = set(DEFAULT_STATUSES)
    for shot in master_data:
        if shot.assign: assign_history.add(shot.assign)
        if shot.status: status_history.add(shot.status)
    
    assign_history = sorted(list(assign_history))
    status_history = sorted(list(status_history))

    batch_layout = ui.VGroup({"Spacing": 10}, [
        ui.Label({"Text": f"Batch Update {len(selected_items)} Shots", "Weight": 0, "Alignment": {"AlignHCenter": True}}),
        ui.Label({"Text": "Only checked fields will be overwritten.", "Weight": 0}),

        # Status
        ui.HGroup({"Spacing": 5, "Weight": 0}, [
            ui.CheckBox({"ID": "chk_status", "Text": "Status", "Weight": 0}),
            ui.LineEdit({"ID": "val_status_line", "Weight": 2}),
            ui.ComboBox({"ID": "val_status_combo", "Weight": 1}),
        ]),

        # Assign
        ui.HGroup({"Spacing": 5, "Weight": 0}, [
            ui.CheckBox({"ID": "chk_assign", "Text": "Assign", "Weight": 0}),
            ui.LineEdit({"ID": "val_assign_line", "Weight": 2}),
            ui.ComboBox({"ID": "val_assign_combo", "Weight": 1}),
        ]),

        # Difficulty
        ui.HGroup({"Spacing": 5, "Weight": 0}, [
            ui.CheckBox({"ID": "chk_diff", "Text": "Difficulty", "Weight": 0}),
            ui.LineEdit({"ID": "val_diff", "Weight": 1}),
        ]),

        # Prior
        ui.HGroup({"Spacing": 5, "Weight": 0}, [
            ui.CheckBox({"ID": "chk_prior", "Text": "Prior", "Weight": 0}),
            ui.LineEdit({"ID": "val_prior", "Weight": 1}),
        ]),

        # Notes
        ui.HGroup({"Spacing": 5, "Weight": 0}, [
            ui.CheckBox({"ID": "chk_notes", "Text": "Notes", "Weight": 0}),
            ui.LineEdit({"ID": "val_notes", "Weight": 1}),
        ]),

        # Timeline
        ui.HGroup({"Spacing": 5, "Weight": 0}, [
            ui.CheckBox({"ID": "chk_timeline", "Text": "Timeline", "Weight": 0}),
            ui.ComboBox({"ID": "val_timeline_combo", "Weight": 1}),
        ]),

        # Offset
        ui.HGroup({"Spacing": 5, "Weight": 0}, [
            ui.CheckBox({"ID": "chk_offset", "Text": "Offset (frames)", "Weight": 0}),
            ui.LineEdit({"ID": "val_offset", "Text": "0", "Weight": 1}),
        ]),

        # VFX Marker Checkbox
        ui.CheckBox({"ID": "chk_create_marker", "Text": "Update VFX Markers", "Checked": True, "Weight": 0}),

        ui.HGroup({"Weight": 0}, [
            ui.Button({"ID": "b_ok", "Text": "Apply Batch Update"}),
            ui.Button({"ID": "b_cancel", "Text": "Cancel"}),
        ])
    ])

    batch_win = disp.AddWindow({"WindowTitle": "Batch Edit", "ID": "BatchDialog", "Geometry": [300, 300, 450, 450]}, batch_layout)
    
    # Setup ComboBoxes
    a_combo = batch_win.Find("val_assign_combo")
    a_combo.AddItem("Recent...")
    for p in assign_history: a_combo.AddItem(p)
    
    s_combo = batch_win.Find("val_status_combo")
    s_combo.AddItem("Select...")
    for s in status_history: s_combo.AddItem(s)

    t_combo = batch_win.Find("val_timeline_combo")
    t_names = []
    projectManager = resolve.GetProjectManager()
    project = projectManager.GetCurrentProject()
    if project:
        count = project.GetTimelineCount()
        for i in range(1, int(count) + 1):
            t_names.append(project.GetTimelineByIndex(i).GetName())
    t_names.sort()
    for name in t_names:
        t_combo.AddItem(name)

    def on_b_assign_select(ev):
        idx = a_combo.CurrentIndex
        if idx > 0: batch_win.Find("val_assign_line").Text = a_combo.GetItemText(idx)
    
    def on_b_status_select(ev):
        idx = s_combo.CurrentIndex
        if idx > 0: batch_win.Find("val_status_line").Text = s_combo.GetItemText(idx)

    batch_win.On["val_assign_combo"].CurrentIndexChanged = on_b_assign_select
    batch_win.On["val_status_combo"].CurrentIndexChanged = on_b_status_select

    def on_batch_ok(ev):
        # Determine what to update
        updates = {}
        if batch_win.Find("chk_status").Checked: updates[6] = batch_win.Find("val_status_line").Text.strip()
        if batch_win.Find("chk_assign").Checked: updates[7] = batch_win.Find("val_assign_line").Text.strip()
        if batch_win.Find("chk_diff").Checked: updates[8] = batch_win.Find("val_diff").Text.strip()
        if batch_win.Find("chk_prior").Checked: updates[9] = batch_win.Find("val_prior").Text.strip()
        if batch_win.Find("chk_notes").Checked: updates[10] = batch_win.Find("val_notes").Text.strip()
        if batch_win.Find("chk_timeline").Checked: updates[11] = batch_win.Find("val_timeline_combo").GetItemText(batch_win.Find("val_timeline_combo").CurrentIndex)

        # Determine if offset is checked
        do_offset = batch_win.Find("chk_offset").Checked
        offset_val = 0
        if do_offset:
            try:
                offset_val = int(batch_win.Find("val_offset").Text.strip())
            except ValueError:
                do_offset = False

        if not updates and not do_offset:
            print("No fields checked; nothing to update.")
            batch_win.Hide()
            return

        # Signature of selected items for batch matching
        to_move_sigs = []
        for it in selected_items.values() if isinstance(selected_items, dict) else selected_items:
            to_move_sigs.append((it.Text[1], it.Text[4])) # Name, StartTC

        # Determine frame rate from active project/timeline
        projectManager = resolve.GetProjectManager()
        project = projectManager.GetCurrentProject()
        current_timeline = project.GetCurrentTimeline() if project else None
        try:
            f_fps = float(current_timeline.GetSetting('timelineFrameRate'))
        except:
            f_fps = 24.0

        # Update Master Data
        for shot in master_data:
            if (shot.name, shot.start_tc) in to_move_sigs:
                if 6 in updates: shot.status = updates[6]
                if 7 in updates: shot.assign = updates[7]
                if 8 in updates: shot.difficulty = updates[8]
                if 9 in updates: shot.prior = updates[9]
                if 10 in updates: shot.notes = updates[10]
                if 11 in updates: shot.timeline = updates[11]
                if do_offset:
                    s_frame = timecode_to_frame(shot.start_tc, f_fps)
                    e_frame = timecode_to_frame(shot.end_tc, f_fps)
                    shot.start_tc = frame_to_timecode(s_frame + offset_val, f_fps)
                    shot.end_tc = frame_to_timecode(e_frame + offset_val, f_fps)

                if batch_win.Find("chk_create_marker").Checked:
                    update_shot_vfx_marker(shot)

        # Update Tree Items in UI
        for it in selected_items.values() if isinstance(selected_items, dict) else selected_items:
            for col, val in updates.items():
                it.SetText(col, val)
            if do_offset:
                s_frame = timecode_to_frame(it.Text[4], f_fps)
                e_frame = timecode_to_frame(it.Text[5], f_fps)
                it.SetText(4, frame_to_timecode(s_frame + offset_val, f_fps))
                it.SetText(5, frame_to_timecode(e_frame + offset_val, f_fps))

        batch_win.Hide()
        # refresh_tree() is bypassed for speed
        auto_save_session()
        print(f"Batch updated {len(to_move_sigs)} items.")

    def on_batch_cancel(ev):
        batch_win.Hide()

    batch_win.On["b_ok"].Clicked = on_batch_ok
    batch_win.On["b_cancel"].Clicked = on_batch_cancel
    batch_win.On.BatchDialog.Close = on_batch_cancel
    batch_win.Show()

def on_batch_edit_clicked(ev):
    selected_items = tree.SelectedItems()
    if not selected_items:
        print("Please select multiple items to batch edit.")
        return
    open_batch_edit_dialog(selected_items)

def on_edit_clicked(ev):
    selected_items = tree.SelectedItems()
    if not selected_items or len(selected_items) != 1:
        print("Please select exactly one item to edit.")
        return
        
    items = selected_items.values() if isinstance(selected_items, dict) else selected_items
    item = next(iter(items))
    open_edit_dialog(item, is_new=False)

def on_delete_clicked(ev):
    selected_items = tree.SelectedItems()
    if not selected_items:
        print("Please select an item to delete.")
        return
        
    to_remove_sigs = []
    items_to_process = selected_items.values() if type(selected_items) is dict else selected_items
    for it in items_to_process:
        to_remove_sigs.append((it.Text[1], it.Text[4])) # Name, StartTC

    global master_data
    master_data = [shot for shot in master_data if (shot.name, shot.start_tc) not in to_remove_sigs]
    refresh_tree()
    auto_save_session()

def on_tree_double_clicked(ev):
    item = ev["item"]
    if not item:
        return
    start_tc = item.Text[4]
    timeline_name = item.Text[11]
    
    projectManager = resolve.GetProjectManager()
    project = projectManager.GetCurrentProject()
    if not project:
        return
        
    current_timeline = project.GetCurrentTimeline()
    
    # Check if we need to switch timelines
    if timeline_name and current_timeline.GetName() != timeline_name:
        found = False
        for i in range(1, int(project.GetTimelineCount()) + 1):
            t = project.GetTimelineByIndex(i)
            if t.GetName() == timeline_name:
                project.SetCurrentTimeline(t)
                current_timeline = t
                found = True
                print(f"Switched to timeline: {timeline_name}")
                break
        if not found:
            print(f"Warning: Timeline '{timeline_name}' not found in project.")

    current_timeline.SetCurrentTimecode(start_tc)
    print(f"Jumped playhead to {start_tc}")

def on_import_csv(ev):
    file_path = request_file(title="Open CSV", mode="open")
    if not file_path:
        return
        
    if not os.path.exists(file_path):
        print(f"File not found: {file_path}")
        return
        
    try:
        global master_data
        with open(file_path, "r", newline="", encoding="utf-8") as f:
            reader = csv.reader(f)
            next(reader)
            for row in reader:
                if not row:
                    continue
                while len(row) < 12:
                    row.append("")
                master_data.append(ShotRecord.from_list(row[:12]))
        refresh_tree()
        auto_save_session()
        print(f"Imported CSV from {file_path}")
    except Exception as e:
        print(f"Failed to import CSV: {e}")

def on_export_csv(ev):
    file_path = request_file(title="Save as CSV", mode="save")
    if not file_path:
        return
        
    dir_path = os.path.dirname(file_path)
    if dir_path and not os.path.exists(dir_path):
        os.makedirs(dir_path)
        
    with open(file_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["Shot ID", "Clip Name", "Thumbnails", "Duration", "Start TC", "End TC", "Status", "Assign", "Difficulty", "Prior", "Notes", "Timeline"])
        
        export_scope = win.Find("export_scope_combo").CurrentIndex
        
        if export_scope == 0:
            # All Items (Export from Master Data)
            for shot in master_data:
                writer.writerow(shot.to_list())
            scope_name = "All"
        elif export_scope == 1:
            # Filtered Items (Export from Tree)
            child_count = tree.TopLevelItemCount()
            for i in range(child_count):
                item = tree.TopLevelItem(i)
                writer.writerow([item.Text[c] for c in range(12)])
            scope_name = "Filtered"
        elif export_scope == 2:
            # Selected Items
            selected_items = tree.SelectedItems()
            if selected_items:
                items_list = selected_items.values() if type(selected_items) is dict else selected_items
                for item in items_list:
                    writer.writerow([item.Text[c] for c in range(12)])
            scope_name = "Selected"
                
    print(f"Exported to {file_path} ({scope_name} items)")

def on_export_excel_clicked(ev):
    # 1. Show column selection dialog
    dialog_layout = ui.VGroup({"Spacing": 10}, [
        ui.Label({"Text": "Select columns to export to Excel:", "Weight": 0}),
        
        ui.HGroup({"Spacing": 10, "Weight": 0}, [
            ui.VGroup({"Weight": 1}, [
                ui.CheckBox({"ID": "col_shot_id", "Text": "Shot ID", "Checked": True}),
                ui.CheckBox({"ID": "col_clip_name", "Text": "Clip Name", "Checked": True}),
                ui.CheckBox({"ID": "col_thumb", "Text": "Thumbnails", "Checked": True}),
                ui.CheckBox({"ID": "col_duration", "Text": "Duration", "Checked": True}),
            ]),
            ui.VGroup({"Weight": 1}, [
                ui.CheckBox({"ID": "col_start_tc", "Text": "Start TC", "Checked": True}),
                ui.CheckBox({"ID": "col_end_tc", "Text": "End TC", "Checked": True}),
                ui.CheckBox({"ID": "col_status", "Text": "Status", "Checked": True}),
                ui.CheckBox({"ID": "col_assign", "Text": "Assign", "Checked": True}),
            ]),
            ui.VGroup({"Weight": 1}, [
                ui.CheckBox({"ID": "col_difficulty", "Text": "Difficulty", "Checked": True}),
                ui.CheckBox({"ID": "col_prior", "Text": "Prior", "Checked": True}),
                ui.CheckBox({"ID": "col_notes", "Text": "Notes", "Checked": True}),
                ui.CheckBox({"ID": "col_timeline", "Text": "Timeline", "Checked": True}),
            ])
        ]),
        
        ui.HGroup({"Weight": 0}, [
            ui.Button({"ID": "exp_ok", "Text": "Export"}),
            ui.Button({"ID": "exp_cancel", "Text": "Cancel"}),
        ])
    ])
    
    dialog = disp.AddWindow({"WindowTitle": "Export to Excel", "ID": "ExcelExportDialog", "Geometry": [300, 300, 450, 220]}, dialog_layout)
    
    def on_cancel(ev):
        dialog.Hide()
        
    dialog.On["exp_cancel"].Clicked = on_cancel
    dialog.On.ExcelExportDialog.Close = on_cancel
    
    def on_ok(ev):
        # 2. Gather selected columns
        columns_map = {
            "col_shot_id": ("Shot ID", 0),
            "col_clip_name": ("Clip Name", 1),
            "col_thumb": ("Thumbnails", 2),
            "col_duration": ("Duration", 3),
            "col_start_tc": ("Start TC", 4),
            "col_end_tc": ("End TC", 5),
            "col_status": ("Status", 6),
            "col_assign": ("Assign", 7),
            "col_difficulty": ("Difficulty", 8),
            "col_prior": ("Prior", 9),
            "col_notes": ("Notes", 10),
            "col_timeline": ("Timeline", 11)
        }
        
        selected_cols = []
        for id_name, col_info in columns_map.items():
            if dialog.Find(id_name).Checked:
                selected_cols.append(col_info) # list of (header_name, index)
                
        dialog.Hide()
        
        if not selected_cols:
            print("No columns selected for export.")
            return
            
        # 3. Prompt user for output location
        file_path = request_file(title="Save Excel File", mode="save", ext_filter="Excel Files (*.xlsx)", default_ext=".xlsx")
        if not file_path:
            return
            
        dir_path = os.path.dirname(file_path)
        if dir_path and not os.path.exists(dir_path):
            os.makedirs(dir_path)
            
        # 4. Gather target data based on Export Scope
        export_scope = win.Find("export_scope_combo").CurrentIndex
        target_shots = []
        if export_scope == 0:
            target_shots = master_data
        elif export_scope == 1:
            child_count = tree.TopLevelItemCount()
            for i in range(child_count):
                item = tree.TopLevelItem(i)
                target_shots.append(ShotRecord.from_list([item.Text[c] for c in range(12)]))
        elif export_scope == 2:
            selected_items = tree.SelectedItems()
            if selected_items:
                items_list = selected_items.values() if type(selected_items) is dict else selected_items
                for item in items_list:
                    target_shots.append(ShotRecord.from_list([item.Text[c] for c in range(12)]))
                    
        # 5. Create Excel Workbook using openpyxl
        try:
            import openpyxl
            from openpyxl.drawing.image import Image
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Shots"
            
            # Write Header row
            headers = [col[0] for col in selected_cols]
            ws.append(headers)
            
            # Configure style/fonts for header
            for col_num in range(1, len(headers) + 1):
                cell = ws.cell(row=1, column=col_num)
                cell.font = openpyxl.styles.Font(bold=True)
                
            # Create a temp folder for downloading thumbnails (if Thumbnail column selected)
            has_thumb = any(col[0] == "Thumbnails" for col in selected_cols)
            temp_dir = None
            if has_thumb:
                temp_dir = os.path.join(dir_path, "_temp_xlsx_thumbs")
                if not os.path.exists(temp_dir):
                    os.makedirs(temp_dir)
                    
            projectManager = resolve.GetProjectManager()
            project = projectManager.GetCurrentProject()
            
            # Populate rows
            for row_idx, shot in enumerate(target_shots, start=2):
                shot_data = shot.to_list()
                
                # Set a tall row height if thumbnails are present
                if has_thumb:
                    ws.row_dimensions[row_idx].height = 55
                    
                for col_idx, (col_name, data_idx) in enumerate(selected_cols, start=1):
                    val = shot_data[data_idx]
                    
                    if col_name == "Thumbnails" and temp_dir and project:
                        # Grab thumbnail and insert
                        shot_id = shot.shot_id if shot.shot_id else f"shot_{row_idx}"
                        start_tc = shot.start_tc
                        timeline_name = shot.timeline
                        
                        if start_tc:
                            # Resolve timeline and switch
                            curr_timeline = project.GetCurrentTimeline()
                            if timeline_name and curr_timeline.GetName() != timeline_name:
                                for i in range(1, int(project.GetTimelineCount()) + 1):
                                    t = project.GetTimelineByIndex(i)
                                    if t.GetName() == timeline_name:
                                        project.SetCurrentTimeline(t)
                                        curr_timeline = t
                                        break
                                        
                            thumb_path = os.path.join(temp_dir, f"{shot_id}.png")
                            success = save_clip_thumbnail_to_file(curr_timeline, start_tc, thumb_path)
                            if success:
                                try:
                                    # Scale down thumbnail to ~120x68 pixels for Excel
                                    resize_image_opencv(thumb_path, 120)
                                    
                                    img = Image(thumb_path)
                                    ws.add_image(img, openpyxl.utils.get_column_letter(col_idx) + str(row_idx))
                                except Exception as img_err:
                                    print(f"Error inserting image to Excel: {img_err}")
                                    ws.cell(row=row_idx, column=col_idx, value="[Image Error]")
                            else:
                                ws.cell(row=row_idx, column=col_idx, value="[No Thumbnail]")
                        else:
                            ws.cell(row=row_idx, column=col_idx, value="")
                    else:
                        ws.cell(row=row_idx, column=col_idx, value=str(val))
                        
            # Adjust column widths (making Thumbnail column wider)
            for col_idx, (col_name, _) in enumerate(selected_cols, start=1):
                col_letter = openpyxl.utils.get_column_letter(col_idx)
                if col_name == "Thumbnails":
                    ws.column_dimensions[col_letter].width = 18
                else:
                    # Auto-fit based on max content length
                    max_len = len(col_name)
                    for r in range(1, len(target_shots) + 2):
                        val_str = str(ws.cell(row=r, column=col_idx).value or "")
                        if val_str and len(val_str) > max_len:
                            max_len = len(val_str)
                    ws.column_dimensions[col_letter].width = min(max_len + 3, 50)
                    
            wb.save(file_path)
            print(f"Excel file saved successfully to {file_path}")
            
            # Clean up temp thumbnail images and folder
            if temp_dir and os.path.exists(temp_dir):
                for f in os.listdir(temp_dir):
                    try:
                        os.remove(os.path.join(temp_dir, f))
                    except:
                        pass
                try:
                    os.rmdir(temp_dir)
                except:
                    pass
                    
            # Switch back to edit page
            if project:
                resolve.OpenPage("edit")
                
        except Exception as xl_err:
            print(f"Failed to export to Excel: {xl_err}")
            
    dialog.On["exp_ok"].Clicked = on_ok
    dialog.Show()

def on_sync_gsheet_clicked(ev):
    """Handles the two-way sync with Google Sheets."""
    global master_data
    gsheet_field = win.Find("gsheet_url")
    url = gsheet_field.Text.strip() if gsheet_field else ""
    if not url:
        print("Error: No Google Sheet Web App URL provided.")
        return
        
    print("Starting Google Sheets Sync...")
    
    # 1. Push Local Data to GSheet
    # We only send 11 columns (excluding Timeline)
    push_data = []
    for shot in master_data:
        push_data.append(shot.to_list()[:11])
        
    try:
        # PUSH (POST)
        json_data = json.dumps(push_data).encode("utf-8")
        req = urllib.request.Request(url, data=json_data, method='POST')
        req.add_header('Content-Type', 'application/json')
        
        with urllib.request.urlopen(req) as response:
            status = response.read().decode('utf-8')
            if "Success" not in status:
                print(f"Push failed: {status}")
                return
            print("Successfully pushed data to Google Sheets.")
            
        # 2. Pull Latest Data from GSheet (GET)
        with urllib.request.urlopen(url) as response:
            pull_json = response.read().decode('utf-8')
            sheet_rows = json.loads(pull_json)
            
            if not sheet_rows or len(sheet_rows) <= 1:
                print("Sheet is empty or has only headers.")
                return
            
            # Headers should be sheet_rows[0]
            # Data starts at sheet_rows[1]
            updated_count = 0
            for row in sheet_rows[1:]:
                if not row: continue
                while len(row) < 11: row.append("")
                
                shot_id = str(row[0]).strip()
                if not shot_id: continue
                
                # Match by Shot ID and update local info
                match_found = False
                for ld in master_data:
                    if ld.shot_id == shot_id:
                        # Update fields: name, thumb, dur, start, end, status, assign, diff, prior, notes
                        # Index mapping: 0:id, 1:name, 2:thumb, 3:dur, 4:start, 5:end, 6:status, 7:assign, 8:diff, 9:prior, 10:notes
                        ld.name = str(row[1])
                        ld.thumb = str(row[2])
                        ld.duration = str(row[3])
                        ld.start_tc = str(row[4])
                        ld.end_tc = str(row[5])
                        ld.status = str(row[6])
                        ld.assign = str(row[7])
                        ld.difficulty = str(row[8])
                        ld.prior = str(row[9])
                        ld.notes = str(row[10])
                        match_found = True
                        updated_count += 1
                        break
                
                if not match_found:
                    # Optional: Add as new shot if not found? 
                    # user said 'append new' in sheet, but pull usually updates existing resolve clips.
                    # If it's a new ID from sheet, we'll keep it as is for now. 
                    # Generally, resolve clips should stay anchored to resolve timelines.
                    pass
            
            print(f"Finished pulling updates. Updated {updated_count} local records.")
            refresh_tree()
            auto_save_session()
            
    except Exception as e:
        print(f"Sync error: {e}")

def on_filter_changed(ev):
    refresh_tree()

def on_link_session_clicked(ev):
    """Allows user to manually link a project to a CSV file."""
    global current_session_file
    file_path = request_file(title="Link Session to CSV", mode="save")
    if not file_path:
        return
        
    proj_name = get_project_name()
    if not proj_name:
        print("Warning: Currently no project is open in Resolve.")
        return
        
    current_session_file = file_path
    win.Find("session_path").Text = file_path
    
    # Save the mapping to Fusion preferences
    sessions = fu.GetData("FxShotManager_Sessions") or {}
    sessions[proj_name] = {
        "path": file_path,
        "seq": global_seq_name,
        "shot": global_shot_num
    }
    fu.SetData("FxShotManager_Sessions", sessions)
    
    # Prompt to load data if it already exists
    if os.path.exists(file_path):
        import_option = request_yes_no(title="Import Data?", message="The file already exists. Should we load the shots from this file now?")
        if import_option:
            # We reuse the logic from on_import_csv but directly using the file_path
            global master_data
            try:
                temp_data = []
                with open(file_path, "r", newline="", encoding="utf-8") as f:
                    reader = csv.reader(f)
                    next(reader)
                    for row in reader:
                        while len(row) < 12:
                            row.append("")
                        temp_data.append(ShotRecord.from_list(row[:12]))
                master_data = temp_data
                refresh_tree()
            except Exception as e:
                print(f"Failed to import existing data: {e}")
    else:
        # Create empty file
        auto_save_session()

    print(f"Linked Project '{proj_name}' to {file_path}")

def on_clear_filter(ev):
    win.Find("filter_text").Text = ""
    refresh_tree()

def on_export_thumbs(ev):
    dir_path = request_file(title="Select Thumbnail Export Folder", mode="directory")
    if not dir_path:
        return
        
    projectManager = resolve.GetProjectManager()
    project = projectManager.GetCurrentProject()
    if not project:
        return
    
    export_scope = win.Find("export_scope_combo").CurrentIndex
    
    # Gather target data
    target_shots = []
    if export_scope == 0:
        target_shots = master_data
    elif export_scope == 1:
        child_count = tree.TopLevelItemCount()
        for i in range(child_count):
            item = tree.TopLevelItem(i)
            target_shots.append(ShotRecord.from_list([item.Text[c] for c in range(12)]))
    elif export_scope == 2:
        selected_items = tree.SelectedItems()
        if selected_items:
            items_list = selected_items.values() if type(selected_items) is dict else selected_items
            for item in items_list:
                target_shots.append(ShotRecord.from_list([item.Text[c] for c in range(12)]))

    # Switch page to Color
    resolve.OpenPage("color")
    
    exported_count = 0
    for shot in target_shots:
        shot_id = shot.shot_id
        start_tc = shot.start_tc
        timeline_name = shot.timeline
        
        if not shot_id:
            continue
            
        # Switch timeline if needed
        curr_timeline = project.GetCurrentTimeline()
        if timeline_name and curr_timeline.GetName() != timeline_name:
            for i in range(1, int(project.GetTimelineCount()) + 1):
                t = project.GetTimelineByIndex(i)
                if t.GetName() == timeline_name:
                    project.SetCurrentTimeline(t)
                    curr_timeline = t
                    print(f"Switched to {timeline_name} for thumbnail capture.")
                    break
            
        target_file = os.path.join(dir_path, f"{shot_id}.png")
        success = save_clip_thumbnail_to_file(curr_timeline, start_tc, target_file)
        if success:
            exported_count += 1
            # Resize image to 720 width as requested by the post-process
            resize_image_opencv(target_file, 720)
            
    print(f"Batch thumbnail export to {dir_path} complete. Exported {exported_count} thumbnails.")
    resolve.OpenPage("edit")

def on_export_clips_clicked(ev):
    projectManager = resolve.GetProjectManager()
    project = projectManager.GetCurrentProject()
    if not project:
        print("No active project.")
        return
        
    dialog_layout = ui.VGroup({"Spacing": 10}, [
        ui.HGroup({"Weight": 0}, [
            ui.VGroup({"Weight": 1}, [
                ui.Label({"Text": "Project Name", "Weight": 0}),
                ui.LineEdit({"ID": "e_proj", "Text": global_proj_name, "Weight": 0}),
            ]),
            ui.VGroup({"Weight": 1}, [
                ui.Label({"Text": "Version", "Weight": 0}),
                ui.LineEdit({"ID": "e_version", "Text": global_version, "Weight": 0}),
            ]),
        ]),
        ui.HGroup({"Weight": 0}, [
            ui.VGroup({"Weight": 2}, [
                ui.Label({"Text": "Render Preset", "Weight": 0}),
                ui.ComboBox({"ID": "e_preset", "Weight": 0}),
            ]),
            ui.VGroup({"Weight": 3}, [
                ui.Label({"Text": "Output Folder", "Weight": 0}),
                ui.HGroup({"Weight": 0}, [
                    ui.LineEdit({"ID": "e_dir", "Text": "", "ReadOnly": True, "Weight": 1}),
                    ui.Button({"ID": "e_browse", "Text": "Browse", "Weight": 0}),
                ]),
            ])
        ]),
        ui.Button({"ID": "e_analyze", "Text": "Analyze Timeline for Clips", "Weight": 0}),
        
        ui.Label({"Text": "Discovered Clips (Select to edit identifier):", "Weight": 0}),
        ui.Tree({
            "ID": "e_tree", 
            "Weight": 1,
            "ColumnCount": 6,
            "AlternatingRowColors": True,
            "SelectionMode": "ExtendedSelection"
        }),
        
        ui.HGroup({"Weight": 0}, [
            ui.Label({"Text": "Selected Identifier:", "Weight": 0}),
            ui.LineEdit({"ID": "e_id_edit", "Weight": 1}),
            ui.Button({"ID": "e_id_apply", "Text": "Apply", "Weight": 0}),
        ]),
        
        ui.HGroup({"Weight": 0}, [
            ui.Button({"ID": "e_start", "Text": "Start Render", "Weight": 0}),
            ui.Button({"ID": "e_cancel", "Text": "Cancel", "Weight": 0}),
        ]),
        ui.VGap(1)
    ])
    
    dialog = disp.AddWindow({"WindowTitle": "Export Clips", "ID": "ExportClipsDialog", "Geometry": [250, 250, 800, 600]}, dialog_layout)
    
    preset_combo = dialog.Find("e_preset")
    preset_combo.AddItem("Current Settings")
    presets = project.GetRenderPresetList()
    presets = sorted(presets)
    
    if presets:
        for p in presets:
            preset_combo.AddItem(p)
    preset_combo.CurrentIndex = 0
            
    clip_tree = dialog.Find("e_tree")
    clip_tree.SetHeaderLabels(["Shot ID", "Track", "Clip Name", "Start TC", "End TC", "Identifier"])
    clip_tree.ColumnWidth[0] = 120
    clip_tree.ColumnWidth[1] = 60
    clip_tree.ColumnWidth[2] = 200
    clip_tree.ColumnWidth[3] = 90
    clip_tree.ColumnWidth[4] = 90
    clip_tree.ColumnWidth[5] = 120
    
    analyzed_clips_data = [] 

    def on_e_browse(ev):
        folder = request_file(title="Select Output Folder", mode="directory")
        if folder: dialog.Find("e_dir").Text = folder
            
    dialog.On["e_browse"].Clicked = on_e_browse
    
    def on_e_analyze(ev):
        clip_tree.Clear()
        analyzed_clips_data.clear()
        
        scope = win.Find("export_scope_combo").CurrentIndex
        target_shots = []
        if scope == 0:
            target_shots = master_data
        elif scope == 1:
            for i in range(tree.TopLevelItemCount()):
                item = tree.TopLevelItem(i)
                target_shots.append(ShotRecord.from_list([item.Text[c] for c in range(12)]))
        elif scope == 2:
            selected_items = tree.SelectedItems()
            if selected_items:
                items_list = selected_items.values() if type(selected_items) is dict else selected_items
                for it in items_list:
                    target_shots.append(ShotRecord.from_list([it.Text[c] for c in range(12)]))
            
        for shot in target_shots:
            if not shot.shot_id or not shot.timeline: continue
            
            curr_timeline = project.GetCurrentTimeline()
            if curr_timeline.GetName() != shot.timeline:
                for i in range(1, int(project.GetTimelineCount()) + 1):
                    t = project.GetTimelineByIndex(i)
                    if t.GetName() == shot.timeline:
                        project.SetCurrentTimeline(t)
                        curr_timeline = t
                        break
            if curr_timeline.GetName() != shot.timeline: continue
            
            try: fps = float(curr_timeline.GetSetting('timelineFrameRate'))
            except: fps = 24.0
            
            target_frame = timecode_to_frame(shot.start_tc, fps)
            
            track_count = curr_timeline.GetTrackCount('video')
            clip_counter = 0
            
            # Loop tracks
            for trk_idx in range(1, track_count + 1):
                if not curr_timeline.GetIsTrackEnabled('video', trk_idx): continue
                
                items = curr_timeline.GetItemListInTrack('video', trk_idx)
                if not items: continue
                
                for item in items:
                    if not item.GetClipEnabled(): continue
                    
                    if item.GetStart() <= target_frame < item.GetEnd():
                        clip_counter += 1
                        identifier = "plate" if clip_counter == 1 else f"plate{clip_counter}"
                        
                        clip_data = {
                            "shot": shot,
                            "track_idx": trk_idx,
                            "item": item,
                            "identifier": identifier,
                            "s_frame": item.GetStart(),
                            "e_frame": item.GetEnd()
                        }
                        analyzed_clips_data.append(clip_data)
                        
                        tree_item = clip_tree.NewItem()
                        tree_item.SetText(0, shot.shot_id)
                        tree_item.SetText(1, f"V{trk_idx}")
                        tree_item.SetText(2, item.GetName())
                        tree_item.SetText(3, frame_to_timecode(item.GetStart(), fps))
                        tree_item.SetText(4, frame_to_timecode(item.GetEnd(), fps))
                        tree_item.SetText(5, identifier)
                        clip_tree.AddTopLevelItem(tree_item)
                        break

        total_missing = len(target_shots) - len(set([c['shot'].shot_id for c in analyzed_clips_data]))
        if total_missing > 0:
            print(f"Skipped {total_missing} shots due to missing clips on timeline.")

    dialog.On["e_analyze"].Clicked = on_e_analyze
    
    def on_e_tree_selection(ev):
        selected = clip_tree.SelectedItems()
        if not selected: return
        sel_list = selected.values() if type(selected) is dict else selected
        if len(sel_list) > 0:
            dialog.Find("e_id_edit").Text = sel_list[0].Text[5]
            
    dialog.On["e_tree"].SelectionChanged = on_e_tree_selection
    
    def on_e_id_apply(ev):
        new_id = dialog.Find("e_id_edit").Text.strip()
        selected = clip_tree.SelectedItems()
        if not selected: return
        sel_list = selected.values() if type(selected) is dict else selected
        for it in sel_list:
            it.SetText(5, new_id)
            for c_data in analyzed_clips_data:
                if c_data['shot'].shot_id == it.Text[0] and f"V{c_data['track_idx']}" == it.Text[1] and c_data['item'].GetName() == it.Text[2]:
                    c_data['identifier'] = new_id
                    
    dialog.On["e_id_apply"].Clicked = on_e_id_apply
    
    def on_e_start(ev):
        out_dir = dialog.Find("e_dir").Text
        if not out_dir:
            print("Please select an output folder.")
            return
            
        if not analyzed_clips_data:
            print("No clips analyzed. Click 'Analyze' first.")
            return
            
        global global_proj_name, global_version
        global_proj_name = dialog.Find("e_proj").Text.strip()
        global_version = dialog.Find("e_version").Text.strip()
        auto_save_session()
        
        preset_name = dialog.Find("e_preset").GetItemText(dialog.Find("e_preset").CurrentIndex)
        if preset_name.lower().strip() not in ("current settings", "current setting"):
            project.LoadRenderPreset(preset_name)
        
        # Group clips by timeline so we don't switch timelines more than needed
        from collections import defaultdict
        timeline_groups = defaultdict(list)
        for c in analyzed_clips_data:
            timeline_groups[c['shot'].timeline].append(c)

        job_ids = []
        cancel_render = False
        
        for tl_name, clips in timeline_groups.items():
            if cancel_render: break
            
            curr_timeline = project.GetCurrentTimeline()
            if curr_timeline.GetName() != tl_name:
                for i in range(1, int(project.GetTimelineCount()) + 1):
                    t = project.GetTimelineByIndex(i)
                    if t.GetName() == tl_name:
                        project.SetCurrentTimeline(t)
                        curr_timeline = t
                        break

            track_count = curr_timeline.GetTrackCount('video')
            
            for c in clips:
                if cancel_render: break
                
                # Isolate the track
                resolve.OpenPage("Edit")
                for trk_idx in range(1, track_count + 1):
                    curr_timeline.SetTrackEnable('video', trk_idx, True)
                for trk_idx in range(1, track_count + 1):
                    if trk_idx != c['track_idx']:
                        curr_timeline.SetTrackEnable('video', trk_idx, False)

                file_name = f"{global_proj_name}_{c['shot'].shot_id}_{c['identifier']}_{global_version}"
                
                settings = {
                    "MarkIn": c['s_frame'],
                    "MarkOut": c['e_frame'] - 1, 
                    "TargetDir": out_dir,
                    "CustomName": file_name
                }
                project.SetRenderSettings(settings)
                job_id = project.AddRenderJob()
                
                if job_id: 
                    job_ids.append(job_id)
                    print(f"Starting render for {file_name} on V{c['track_idx']}...")
                    try:
                        project.StartRendering([job_id])
                        while project.IsRenderingInProgress():
                            time.sleep(0.5)
                            
                        # Check status to see if user manually cancelled
                        status = project.GetRenderJobStatus(job_id)
                        if status and status.get("JobStatus") in ("Cancelled", "Failed"):
                            print(f"Render {status.get('JobStatus').lower()}. Stopping batch...")
                            cancel_render = True
                    except Exception as ex:
                        print(f"Failed to process render: {ex}")
                
            # Restore tracks
            for trk_idx in range(1, track_count + 1):
                curr_timeline.SetTrackEnable('video', trk_idx, True)

        if not cancel_render:
            print(f"Batch rendered {len(job_ids)} clips successfully.")
            
        dialog.Hide()
        
    dialog.On["e_start"].Clicked = on_e_start
    dialog.On["e_cancel"].Clicked = lambda ev: dialog.Hide()
    dialog.On.ExportClipsDialog.Close = lambda ev: dialog.Hide()
    
    dialog.Show()

win.On["export_clips_btn"].Clicked = on_export_clips_clicked
win.On["edit_btn"].Clicked = on_edit_clicked
win.On["batch_edit_btn"].Clicked = on_batch_edit_clicked
win.On["delete_btn"].Clicked = on_delete_clicked
win.On["refresh_btn"].Clicked = add_current_clip
win.On["import_csv_btn"].Clicked = on_import_csv
win.On["export_csv_btn"].Clicked = on_export_csv
win.On["export_excel_btn"].Clicked = on_export_excel_clicked
win.On["export_thumb_btn"].Clicked = on_export_thumbs
win.On["clips_tree"].ItemDoubleClicked = on_tree_double_clicked
win.On["filter_text"].TextChanged = on_filter_changed
win.On["filter_col"].CurrentIndexChanged = on_filter_changed
win.On["clear_filter_btn"].Clicked = on_clear_filter
win.On["link_session_btn"].Clicked = on_link_session_clicked
if win.Find("sync_gsheet_btn"):
    win.On["sync_gsheet_btn"].Clicked = on_sync_gsheet_clicked
win.On.MainWindow.Close = lambda ev: (tk_root.destroy(), disp.ExitLoop(), win.Hide())

# -------------------------------------------------------------
# AUTO-LOAD SESSION
# -------------------------------------------------------------
def load_startup_session():
    global master_data, current_session_file, global_seq_name, global_shot_num
    proj_name = get_project_name()
    if not proj_name:
        return
        
    sessions = fu.GetData("FxShotManager_Sessions") or {}
    if proj_name in sessions:
        data = sessions[proj_name]
        
        # Handle backward compatibility (if it was just a string path before)
        file_path = data if isinstance(data, str) else data.get("path")
        if not file_path:
            return

        if os.path.exists(file_path):
            current_session_file = file_path
            win.Find("session_path").Text = file_path
            
            # Load Global Variables
            if isinstance(data, dict):
                global_seq_name = data.get("seq", "SEQ")
                global_shot_num = int(data.get("shot", 10))
                global_proj_name = data.get("gProj", "PROJ")
                global_version = data.get("gVers", "v01")
                # win.Find("gsheet_url").Text = data.get("gsheet", "")
            
            try:
                with open(file_path, "r", newline="", encoding="utf-8") as f:
                    reader = csv.reader(f)
                    next(reader)
                    for row in reader:
                        master_data.append(ShotRecord.from_list(row))
                refresh_tree()
                print(f"Auto-loaded project session from {file_path}")
            except Exception as e:
                print(f"Auto-load failed: {e}")
        else:
            print(f"Linked session file not found: {file_path}")

load_startup_session()

win.Show()
disp.RunLoop()
win.Hide()
